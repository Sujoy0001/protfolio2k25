from django.shortcuts import render, redirect
from .models import AboutMe, Project, All_Skills, experience, extra_skills
from django.contrib import messages
from django.shortcuts import get_object_or_404
from django.urls import reverse
from openpyxl import Workbook, load_workbook
import os
from django.shortcuts import render



# Create your views here.
def example(request):
    return render(request, 'example.html')

def index(request):
    about_info2 = AboutMe.objects.first() 
    Projects = Project.objects.all().reverse()
    exprate = experience.objects.all()
    skills = extra_skills.objects.all()
    total_projects = Projects.count()
    total_experience = exprate.count()
    return render(request, 'index.html', { 'aboutinfo': about_info2, 'projects': Projects, 'exprate': exprate, 'skills': skills, 'total_projects': total_projects, 'total_experience': total_experience })

def projects(request):
    project_list = Project.objects.all()
    categories = Project.objects.values_list('category', flat=True).distinct()
    
    
    context = {
        'projects': project_list,
        'categories': categories,
    }
    
    return render(request, 'project.html', context)


def project_detail(request, pk):
    project = get_object_or_404(Project, pk=pk)
    return render(request, 'project_detils.html', {'project': project})


def skills(request):
    skill_list = All_Skills.objects.all()
    return render(request, 'skills.html', {'skills': skill_list})

# def contact(request):
#     detills = AboutMe.objects.first()
#     if request.method == "POST":
#         name = request.POST.get("name")
#         email = request.POST.get("email")
#         subject = request.POST.get("subject")
#         message = request.POST.get("message")

#         # Excel file path
#         file_path = "contact_data.xlsx"

#         if os.path.exists(file_path):
#             wb = load_workbook(file_path)
#             ws = wb.active
#         else:
#             wb = Workbook()
#             ws = wb.active
#             ws.append(["Name", "Email", "Subject", "Message"])  # header

#         # Append data
#         ws.append([name, email, subject, message])
#         wb.save(file_path)

#         messages.success(request, "Your message has been sent successfully!")
#         return redirect("contact")  # Replace with your actual URL name

#     return render(request, "contact.html", {'detills': detills})

def review(request):
    # Get all work experience entries ordered by start date (newest first)
    experiences = experience.objects.all()
    
    return render(request, 'reviews.html', {experiences: experiences})



from django.core.mail import send_mail
from django.shortcuts import render
from django.http import HttpResponse

def contact(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        subject = request.POST.get('subject', '')
        message = request.POST.get('message')
        
        try:
            # send email or store data
            messages.success(request, "Your message has been sent successfully!")
        except Exception as e:
            messages.error(request, f"Something went wrong: {e}")

        # Compose the email body
        email_subject = f"New Contact Form Submission: {subject}"
        email_message = f"""
        You have received a new message from {name} ({email}):

        Subject: {subject}

        Message:
        {message}
        """

        # Send the email
        send_mail(
            email_subject,
            email_message,
            'sujoycode999@gmail.com',  # From email address (your email)
            ['recipient-email@example.com'],  # To email address
            fail_silently=False,
        )

        return render(request, 'contact.html', {'success': True})

    return render(request, 'contact.html')

def review(request):
    # Get all work experience entries ordered by start date (newest first)
    experiences = experience.objects.all()
    
    return render(request, 'reviews.html', {'reviews': experiences})

